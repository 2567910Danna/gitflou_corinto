from openpyxl import load_workbook
from datetime import datetime

rut=r'crudDatos.xlsx'

def agregar(ruta:int, datos:dict):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['tareas']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row+1)]
  hoja=archivoExcel.active

  nombre=2
  descripcion=3
  estado=4
  fechaInicio=5
  fechaFinalizado=6
  for i in hojaDatos:
    if not(isinstance(i[0].value,int)):
      identificador=i[0].row
      hoja.cell(row=identificador,column=1).value=identificador-1
      hoja.cell(row=identificador,column=nombre).value=datos['nombre']
      hoja.cell(row=identificador,column=descripcion).value=datos['descripcion']
      hoja.cell(row=identificador,column=estado).value=datos['estado']
      hoja.cell(row=identificador,column=fechaInicio).value=datos['fecha inicio']
      hoja.cell(row=identificador,column=fechaFinalizado).value=datos['fecha finalizacion']
      break
  archivoExcel.save(ruta)
  return